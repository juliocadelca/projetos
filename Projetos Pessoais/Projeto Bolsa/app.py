import openpyxl
import yfinance as yf

# Definir as ações e seus respectivos símbolos
acoes = {
    "BBAS3.SA": "BBAS3",
    "GOAU4.SA": "GOAU4",
    "KLBN4.SA": "KLBN4",
    "ITSA4.SA": "ITSA4",
    "CMIG4.SA": "CMIG4",
    "SAPR4.SA": "SAPR4",
    "TAEE4.SA": "TAEE4",
    "SANB4.SA": "SANB4",
}

# Usar planilha existente
planilha = openpyxl.load_workbook('Carteira de Ações e FIIs.xlsx')
pagina = planilha['ATZ']

# Pegar o preço atualizado de cada ação
linha = 3
for acao, nome in acoes.items():
    ticker = yf.Ticker(acao)
    preco = ticker.history().iloc[-1]['Close']
    pagina.cell(row=linha, column=7).value = nome
    pagina.cell(row=linha, column=8).value = round(preco, 2)
    linha += 1

# Formatar as células para exibirem o preço com duas casas decimais
for col in range(6, 8):
    for row in range(3, linha):
        pagina.cell(row=row, column=col).number_format = '0.00'

# Definir os FIIs e seus respectivos símbolos
fiis = {
    "APTO11.SA": "APTO11",
    "BTLG11.SA": "BTLG11",
    "CPTS11.SA": "CPTS11",
    "GARE11.SA": "GARE11",
    "VGIR11.SA": "VGIR11",
    "VIUR11.SA": "VIUR11",
    "XPSF11.SA": "XPSF11"
}
# Usar planilha existente
pagina1 = planilha['FIIs']

# Pegar o preço atualizado de cada ação
linha = 3
for fii, nome in fiis.items():
    ticker = yf.Ticker(fii)
    preco = ticker.history().iloc[-1]['Close']
    pagina1.cell(row=linha, column=6).value = nome
    pagina1.cell(row=linha, column=7).value = round(preco, 2)
    linha += 1

# Formatar as células para exibirem o preço com duas casas decimais
for col in range(6, 8):
    for row in range(3, linha):
        pagina1.cell(row=row, column=col).number_format = '0.00'

# Salvar a planilha no local especificado
caminho = 'C:\\Users\\TI-JULIO\\Desktop\\Projeto carteira\\Carteira de Ações e FIIs.xlsx'
planilha.save(caminho)
